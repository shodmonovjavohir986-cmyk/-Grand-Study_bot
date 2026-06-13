import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_COLOR = "1E3A5F"
ALT_ROW_COLOR = "EBF3FB"

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style(cell):
    cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")
    cell.border = _border()

def _title_style(cell, size=14):
    cell.font = Font(bold=True, size=size, color=HEADER_COLOR)
    cell.alignment = Alignment(horizontal="center")

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def generate_excel_enrollments(course: dict, enrollments: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ro'yxat"

    # Title
    ws.merge_cells("A1:E1")
    _title_style(ws["A1"])
    ws["A1"] = f"Grand Study — {course['name']} kursi ro'yxati"

    ws.merge_cells("A2:E2")
    ws["A2"] = (f"Narx: {course['price']}  |  Davomiyligi: {course['duration']}  |  "
                f"Sana: {datetime.now().strftime('%Y-%m-%d')}")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="666666")

    # Headers
    headers = ["#", "Ism-Familiya", "Telefon", "Username", "Yozilgan sana"]
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=4, column=col, value=h))

    # Rows
    for i, e in enumerate(enrollments, 1):
        fill = PatternFill("solid", fgColor=ALT_ROW_COLOR if i % 2 == 0 else "FFFFFF")
        row_data = [i, e[0], e[1], f"@{e[2]}" if e[2] else "—", e[3]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 4, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "left")
            cell.border = _border()
            cell.fill = fill

    _set_col_widths(ws, {"A": 5, "B": 28, "C": 18, "D": 18, "E": 20})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_excel_users(users: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    # Title
    ws.merge_cells("A1:F1")
    _title_style(ws["A1"])
    ws["A1"] = f"Grand Study — Barcha foydalanuvchilar  ({datetime.now().strftime('%Y-%m-%d')})"

    # Headers
    headers = ["#", "Ism-Familiya", "Telefon", "Username", "Referal kodi", "Ro'yxat sanasi"]
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=3, column=col, value=h))

    # Rows
    for i, u in enumerate(users, 1):
        fill = PatternFill("solid", fgColor=ALT_ROW_COLOR if i % 2 == 0 else "FFFFFF")
        row_data = [i, u[2], u[3], f"@{u[8]}" if u[8] else "—", u[4], u[7]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 3, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left")
            cell.border = _border()
            cell.fill = fill

    _set_col_widths(ws, {"A": 5, "B": 28, "C": 18, "D": 18, "E": 15, "F": 20})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
